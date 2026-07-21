"""Core spreadsheet metadata sources."""

from __future__ import annotations

__all__ = ["ExcelMetadataSource", "SpreadsheetMetadataSource"]

import abc
import importlib.util
import math
import numbers
import pathlib
import re
import threading
import typing
from types import MappingProxyType

import erlab

if typing.TYPE_CHECKING:
    from collections.abc import Callable, Iterable, Mapping
    from typing import Any

    import xarray as xr


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and value == "")


def _normalize_column_name(value: object, *, description: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{description} must be a non-empty string")
    return value.strip()


def _normalize_mapping(
    mapping: Mapping[str, str] | None, *, description: str
) -> dict[str, str]:
    if mapping is None:
        return {}

    normalized: dict[str, str] = {}
    for source, destination in mapping.items():
        source_name = _normalize_column_name(
            source, description=f"{description} source column"
        )
        destination_name = _normalize_column_name(
            destination, description=f"{description} destination"
        )
        if source_name in normalized:
            raise ValueError(f"Duplicate {description} source column {source_name!r}")
        normalized[source_name] = destination_name
    return normalized


_SIMPLE_FILE_NAME_PATTERN = re.compile(
    r"[A-Za-z0-9_]*?(?P<number>\d+)(?:\.[A-Za-z][A-Za-z0-9]*)*\Z"
)
_INTEGER_PATTERN = re.compile(r"[+-]?(?:0|[1-9]\d*)\Z")
_FLOAT_PATTERN = re.compile(
    r"[+-]?(?:(?:\d+\.\d*|\.\d+)(?:[eE][+-]?\d+)?|\d+[eE][+-]?\d+)\Z"
)
_SIGNED_DIGITS_PATTERN = re.compile(r"[+-]?\d+\Z")
_FILE_RANGE_SEPARATOR_PATTERN = re.compile("[~\N{EN DASH}]")


class _SpreadsheetRow(typing.NamedTuple):
    spreadsheet_row: int
    file_name: object
    values: tuple[typing.Any, ...]


class _FileNumberRange(typing.NamedTuple):
    start: int
    end: int
    row: _SpreadsheetRow


class _SpreadsheetIndex(typing.NamedTuple):
    file_names: dict[str, list[_SpreadsheetRow]]
    ranges: tuple[_FileNumberRange, ...]


class _SpreadsheetMatch(typing.NamedTuple):
    row: _SpreadsheetRow
    number_range: tuple[int, int] | None = None


def _file_number_parse_error(
    value: object, reason: str, *, spreadsheet_row: int | None
) -> ValueError:
    if spreadsheet_row is None:
        location = f"File-number value {value!r}"
    else:
        location = f"File-name cell {value!r} in spreadsheet row {spreadsheet_row}"
    return ValueError(f"{location} was rejected: {reason}")


def _normalize_file_number(value: object, *, spreadsheet_row: int | None = None) -> int:
    if isinstance(value, bool):
        raise _file_number_parse_error(
            value,
            "boolean values are not valid file numbers",
            spreadsheet_row=spreadsheet_row,
        )

    if isinstance(value, numbers.Integral):
        number = int(value)
    elif isinstance(value, numbers.Real):
        numeric = float(value)
        if not math.isfinite(numeric):
            raise _file_number_parse_error(
                value,
                "numeric values must be finite",
                spreadsheet_row=spreadsheet_row,
            )
        if not numeric.is_integer():
            raise _file_number_parse_error(
                value,
                "numeric values must be integers",
                spreadsheet_row=spreadsheet_row,
            )
        number = int(numeric)
    elif isinstance(value, str):
        text = value.strip()
        if not text:
            raise _file_number_parse_error(
                value,
                "the cell is empty after trimming whitespace",
                spreadsheet_row=spreadsheet_row,
            )
        if _SIGNED_DIGITS_PATTERN.fullmatch(text):
            numeric_text = text
        else:
            filename = re.split(r"[/\\]", text)[-1]
            match = _SIMPLE_FILE_NAME_PATTERN.fullmatch(filename)
            if match is None:
                raise _file_number_parse_error(
                    value,
                    "expected decimal digits or a simple filename ending in digits, "
                    "such as 'f_0001.pxt'",
                    spreadsheet_row=spreadsheet_row,
                )
            numeric_text = match.group("number")
        try:
            number = int(numeric_text)
        except ValueError as exc:
            raise _file_number_parse_error(
                value,
                "the numeric portion is too long to convert to an integer",
                spreadsheet_row=spreadsheet_row,
            ) from exc
    else:
        raise _file_number_parse_error(
            value,
            f"unsupported value type {type(value).__name__!r}; expected an integer or "
            "string",
            spreadsheet_row=spreadsheet_row,
        )

    if number < 0:
        raise _file_number_parse_error(
            value,
            "negative file numbers are not allowed",
            spreadsheet_row=spreadsheet_row,
        )
    return number


def _normalize_file_number_range(
    value: object, *, spreadsheet_row: int | None = None
) -> tuple[int, int] | None:
    if not isinstance(value, str):
        return None

    text = value.strip()
    separators = list(_FILE_RANGE_SEPARATOR_PATTERN.finditer(text))
    if not separators:
        return None
    if len(separators) != 1:
        raise _file_number_parse_error(
            value,
            "file ranges must contain exactly one '~' or en dash separator",
            spreadsheet_row=spreadsheet_row,
        )

    separator = separators[0]
    start_value = text[: separator.start()].strip()
    end_value = text[separator.end() :].strip()
    if not start_value or not end_value:
        raise _file_number_parse_error(
            value,
            "file ranges must have a value on both sides of the separator",
            spreadsheet_row=spreadsheet_row,
        )

    try:
        start = _normalize_file_number(start_value)
    except ValueError as exc:
        raise _file_number_parse_error(
            value,
            f"range start {start_value!r} is not a valid file number or file name",
            spreadsheet_row=spreadsheet_row,
        ) from exc
    try:
        end = _normalize_file_number(end_value)
    except ValueError as exc:
        raise _file_number_parse_error(
            value,
            f"range end {end_value!r} is not a valid file number or file name",
            spreadsheet_row=spreadsheet_row,
        ) from exc
    if start > end:
        raise _file_number_parse_error(
            value,
            f"range start {start} is greater than range end {end}",
            spreadsheet_row=spreadsheet_row,
        )
    return start, end


def _parse_spreadsheet_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value

    text = value.strip()
    if text == "TRUE":
        return True
    if text == "FALSE":
        return False
    if _INTEGER_PATTERN.fullmatch(text):
        return int(text)
    if _FLOAT_PATTERN.fullmatch(text):
        return float(text)
    return value


_SKIP_VALUE = object()


def _coerce_numeric_replacement(
    value: Any, existing: Any, *, kind: str, name: str
) -> Any:
    value = _parse_spreadsheet_value(value)
    if isinstance(existing, bool) or not isinstance(existing, numbers.Number):
        return value
    if isinstance(value, numbers.Number) and not isinstance(value, bool):
        return value
    if not isinstance(value, bool):
        try:
            return float(value)
        except (TypeError, ValueError, OverflowError):
            pass
    erlab.utils.misc.emit_user_level_warning(
        f"Could not convert spreadsheet metadata value {value!r} for "
        f"numeric {kind} {name!r}; the value was skipped"
    )
    return _SKIP_VALUE


def _normalize_row_range(value: object) -> tuple[int, int]:
    if (
        not isinstance(value, tuple)
        or len(value) != 2
        or any(isinstance(row, bool) for row in value)
        or any(not isinstance(row, int) for row in value)
    ):
        raise TypeError("row_range must be a pair of integer row numbers")
    if value[0] < 2 or value[1] < value[0]:
        raise ValueError(
            "row_range must be an increasing inclusive range starting at row 2"
        )
    return value


def _normalize_sheet_name(value: object) -> str | int:
    if isinstance(value, bool) or not isinstance(value, (str, int)):
        raise TypeError("sheet_name must be a worksheet name or zero-based index")
    if isinstance(value, int) and value < 0:
        raise ValueError("sheet_name indices must be non-negative")
    if isinstance(value, str) and not value:
        raise ValueError("sheet_name must not be empty")
    return value


class _SpreadsheetMetadataValues:
    """Resolved values for one spreadsheet row."""

    def __init__(
        self,
        coordinate_values: Mapping[str, Any],
        attribute_values: Mapping[str, Any],
        *,
        overwrite: bool,
    ) -> None:
        self.coordinate_values = dict(coordinate_values)
        self.attribute_values = dict(attribute_values)
        self.overwrite = overwrite

    def apply(
        self, data: xr.DataArray | xr.Dataset | xr.DataTree
    ) -> xr.DataArray | xr.Dataset | xr.DataTree:
        """Apply the resolved values to every leaf data array."""
        return erlab.utils.array.apply_dataarray_func(data, self._apply_dataarray)

    def _apply_dataarray(self, data: xr.DataArray) -> xr.DataArray:
        result = data
        for name, value in self.coordinate_values.items():
            if name in result.dims:
                if self.overwrite:
                    raise ValueError(
                        f"Spreadsheet metadata cannot overwrite dimension coordinate "
                        f"{name!r}"
                    )
                continue
            if name in result.coords:
                if not self.overwrite:
                    continue
                if result[name].ndim != 0:
                    raise ValueError(
                        f"Spreadsheet metadata cannot overwrite non-scalar "
                        f"coordinate {name!r}"
                    )
                value = _coerce_numeric_replacement(
                    value,
                    result[name].item(),
                    kind="coordinate",
                    name=name,
                )
                if value is _SKIP_VALUE:
                    continue
            else:
                value = _parse_spreadsheet_value(value)
            try:
                result = result.assign_coords({name: value})
            except (TypeError, ValueError, OverflowError):
                erlab.utils.misc.emit_user_level_warning(
                    f"Could not apply spreadsheet metadata value {value!r} to "
                    f"coordinate {name!r}; the value was skipped"
                )

        for name, value in self.attribute_values.items():
            if name in result.attrs:
                if not self.overwrite:
                    continue
                value = _coerce_numeric_replacement(
                    value,
                    result.attrs[name],
                    kind="attribute",
                    name=name,
                )
                if value is _SKIP_VALUE:
                    continue
            else:
                value = _parse_spreadsheet_value(value)
            result = result.assign_attrs({name: value})
        return result


class SpreadsheetMetadataSource(abc.ABC):
    """Base class for spreadsheet-backed loader metadata.

    Parameters
    ----------
    file_name_column
        Header of the column containing the file name or number used to match loaded
        data. For path-based loads, whitespace-trimmed cell text is first matched
        exactly and case-sensitively against the file name without its path or final
        extension. The loader derives a file number from the path with
        :meth:`erlab.io.dataloader.LoaderBase.infer_index` for numeric fallback and to
        check whether a range also claims a direct match. That number is compared with
        spreadsheet cells containing an integer or a simple filename ending in a number,
        such as ``f_0001.pxt``. Cells can also contain an inclusive range separated by
        ``~`` or an en dash, such as ``f_0015~20`` or ``f_0015~f_0020``.
        `file_name_column` may be omitted while inspecting :meth:`get_sheet_names` or
        :meth:`get_column_names`.
    coordinate_mapping
        Mapping from spreadsheet headers to scalar coordinate names. This and
        `attribute_mapping` may both be omitted while inspecting
        :meth:`get_column_names`, but at least one is required when applying metadata.
        The normalized mapping is immutable after construction.
    attribute_mapping
        Mapping from spreadsheet headers to attribute names. The normalized mapping is
        immutable after construction.
    sheet_name
        Worksheet name or zero-based worksheet index. If omitted, the source chooses its
        default worksheet.
    overwrite
        Whether spreadsheet values replace existing scalar coordinates and attributes.
    row_range
        Optional inclusive range of 1-based data row numbers to search. Row 1 contains
        headers, so the first valid data row is 2.

    Notes
    -----
    Configured source-column names are matched exactly first. If no exact header exists,
    a spreadsheet header containing newlines also matches the same name with each
    newline replaced by a space.

    .. versionadded:: 3.25.0
    """

    _cache_row_index = True

    def __init__(
        self,
        *,
        file_name_column: str | None = None,
        coordinate_mapping: Mapping[str, str] | None = None,
        attribute_mapping: Mapping[str, str] | None = None,
        sheet_name: str | int | None = None,
        overwrite: bool = False,
        row_range: tuple[int, int] | None = None,
    ) -> None:
        self._file_name_column = (
            None
            if file_name_column is None
            else _normalize_column_name(
                file_name_column, description="file name column"
            )
        )
        self._coordinate_mapping = MappingProxyType(
            _normalize_mapping(coordinate_mapping, description="coordinate mapping")
        )
        self._attribute_mapping = MappingProxyType(
            _normalize_mapping(attribute_mapping, description="attribute mapping")
        )
        overlapping_destinations = set(self._coordinate_mapping.values()) & set(
            self._attribute_mapping.values()
        )
        if overlapping_destinations:
            names = ", ".join(repr(v) for v in sorted(overlapping_destinations))
            raise ValueError(
                "Coordinate and attribute mappings have overlapping destinations: "
                f"{names}"
            )

        if not isinstance(overwrite, bool):
            raise TypeError("overwrite must be a boolean")
        self._sheet_name = (
            None if sheet_name is None else _normalize_sheet_name(sheet_name)
        )
        self._overwrite = overwrite
        self._row_range = None if row_range is None else _normalize_row_range(row_range)
        self._lock = threading.RLock()
        self._row_index: _SpreadsheetIndex | None = None
        self._headers: tuple[str, ...] | None = None
        self._sheet_names: tuple[str, ...] | None = None

    @property
    def file_name_column(self) -> str | None:
        """Spreadsheet column used to resolve the file name or number."""
        return self._file_name_column

    @property
    def coordinate_mapping(self) -> Mapping[str, str]:
        """Immutable mapping from spreadsheet columns to coordinate names."""
        return self._coordinate_mapping

    @property
    def attribute_mapping(self) -> Mapping[str, str]:
        """Immutable mapping from spreadsheet columns to attribute names."""
        return self._attribute_mapping

    @property
    def sheet_name(self) -> str | int | None:
        """Selected worksheet name or index."""
        return self._sheet_name

    @property
    def overwrite(self) -> bool:
        """Whether mapped values replace existing metadata."""
        return self._overwrite

    @property
    def row_range(self) -> tuple[int, int] | None:
        """Inclusive range of 1-based spreadsheet data rows to search."""
        return self._row_range

    @property
    @abc.abstractmethod
    def source_name(self) -> str:
        """Human-readable source name used in messages."""

    @abc.abstractmethod
    def _read_rows(self) -> list[list[Any]]:
        """Read the header and configured data rows.

        The first returned row must be the spreadsheet header. If `row_range` is set,
        the remaining rows must begin at its first row and end at its last row.
        """

    def _read_header(self) -> list[Any]:
        """Read only the spreadsheet header row."""
        rows = self._read_rows()
        if not rows:
            return []
        return rows[0]

    def _read_sheet_names(self) -> list[str]:
        """Read worksheet names in workbook order."""
        raise NotImplementedError(
            f"{type(self).__name__} does not support worksheet discovery"
        )

    def _clear_source_cache(self) -> None:  # noqa: B027
        """Clear source-specific cached values."""

    def refresh(self) -> None:
        """Clear and revalidate cached spreadsheet structure.

        Sources that always read current row content only revalidate their cached
        worksheet and header information.
        """
        with self._lock:
            previous = self._headers, self._row_index, self._sheet_names
            self._headers = None
            self._row_index = None
            self._sheet_names = None
            self._clear_source_cache()
            try:
                if self.file_name_column is None or not self._cache_row_index:
                    self._headers = self._normalize_headers(self._read_header())
                    if self.file_name_column is not None:
                        self._resolve_file_name_column(self._headers)
                        self._validate_mapping_columns(self._headers)
                else:
                    self._headers, self._row_index = self._build_index(
                        self._read_rows()
                    )
                    self._validate_mapping_columns(self._headers)
            except Exception:
                self._headers, self._row_index, self._sheet_names = previous
                raise

    def get_sheet_names(self) -> list[str]:
        """Return the worksheet names in workbook order.

        The names are read on the first call and reused until :meth:`refresh` is called.
        A new list is returned on every call.

        Returns
        -------
        list of str
            Visible worksheet names in workbook order.
        """
        with self._lock:
            if self._sheet_names is None:
                names = tuple(self._read_sheet_names())
                if not names:
                    raise ValueError(f"{self.source_name} has no worksheets")
                if any(not name for name in names):
                    raise ValueError(
                        f"{self.source_name} returned an invalid worksheet name"
                    )
                if len(set(names)) != len(names):
                    raise ValueError(
                        f"{self.source_name} has duplicate worksheet names"
                    )
                self._sheet_names = names
            return list(self._sheet_names)

    def get_column_names(self) -> list[str]:
        """Return the spreadsheet column names.

        Only the header row is read on the first call. The result is reused until
        :meth:`refresh` is called. A new list is returned on every call so modifying it
        does not affect the cached column names.

        Returns
        -------
        list of str
            Column names from the first spreadsheet row, in their original order.
        """
        with self._lock:
            if self._headers is None:
                self._headers = self._normalize_headers(self._read_header())
            return list(self._headers)

    def _get_cached_snapshot(
        self,
    ) -> tuple[tuple[str, ...], _SpreadsheetIndex]:
        with self._lock:
            if not self._cache_row_index:
                headers, row_index = self._build_index(self._read_rows())
                if self._headers is None:
                    self._headers = headers
                return headers, row_index
            if self._row_index is None:
                self._headers, self._row_index = self._build_index(self._read_rows())
            if self._headers is None:  # pragma: no cover - assigned with the row index
                raise RuntimeError("Spreadsheet metadata cache is inconsistent")
            return self._headers, self._row_index

    def _metadata_for_file_number(
        self, file_number: object
    ) -> _SpreadsheetMetadataValues | None:
        """Resolve metadata by parsed file number."""
        self._validate_lookup_configuration()
        number = _normalize_file_number(file_number)
        headers, row_index = self._get_cached_snapshot()
        return self._metadata_for_number(number, headers, row_index)

    def _metadata_for_file_name(
        self,
        file_name: str,
        infer_file_number: Callable[[], object | None],
    ) -> tuple[_SpreadsheetMetadataValues | None, str | int | None]:
        """Match a bare file name exactly, parsing numbers only after a miss."""
        self._validate_lookup_configuration()
        headers, row_index = self._get_cached_snapshot()
        direct_rows = row_index.file_names.get(file_name, [])
        if direct_rows:
            direct_matches = [_SpreadsheetMatch(row) for row in direct_rows]
            if len(direct_matches) > 1:
                return (
                    self._metadata_for_matches(file_name, direct_matches, headers),
                    file_name,
                )

            range_number: int | None = None
            if row_index.ranges:
                inferred = infer_file_number()
                if inferred is not None:
                    range_number = _normalize_file_number(inferred)
                else:
                    try:
                        range_number = _normalize_file_number(file_name)
                    except ValueError:
                        range_number = None
            if range_number is not None:
                range_matches = self._matching_ranges(range_number, row_index)
                if range_matches:
                    return (
                        self._metadata_for_matches(
                            range_number,
                            [*direct_matches, *range_matches],
                            headers,
                        ),
                        file_name,
                    )
            return (
                self._metadata_for_matches(file_name, direct_matches, headers),
                file_name,
            )

        inferred = infer_file_number()
        if inferred is None:
            return None, None
        number = _normalize_file_number(inferred)
        return self._metadata_for_number(number, headers, row_index), number

    def _validate_lookup_configuration(self) -> None:
        if not self.coordinate_mapping and not self.attribute_mapping:
            raise ValueError(
                "At least one coordinate or attribute mapping must be provided to "
                "apply spreadsheet metadata"
            )
        if self.file_name_column is None:
            raise ValueError(
                "file_name_column must be provided to apply spreadsheet metadata"
            )

    def _metadata_for_number(
        self,
        number: int,
        headers: tuple[str, ...],
        row_index: _SpreadsheetIndex,
    ) -> _SpreadsheetMetadataValues | None:
        matches: list[_SpreadsheetMatch] = []
        for candidates in row_index.file_names.values():
            for row in candidates:
                try:
                    candidate_number = _normalize_file_number(
                        row.file_name, spreadsheet_row=row.spreadsheet_row
                    )
                except ValueError:
                    continue
                if candidate_number == number:
                    matches.append(_SpreadsheetMatch(row))
        matches.extend(self._matching_ranges(number, row_index))
        return self._metadata_for_matches(number, matches, headers)

    @staticmethod
    def _matching_ranges(
        number: int, row_index: _SpreadsheetIndex
    ) -> list[_SpreadsheetMatch]:
        return [
            _SpreadsheetMatch(item.row, (item.start, item.end))
            for item in row_index.ranges
            if item.start <= number <= item.end
        ]

    def _metadata_for_matches(
        self,
        key: str | int,
        matches: list[_SpreadsheetMatch],
        headers: tuple[str, ...],
    ) -> _SpreadsheetMetadataValues | None:
        resolved_columns = self._resolve_mapping_columns(headers)
        if not matches:
            return None
        matches = sorted(matches, key=lambda match: match.row.spreadsheet_row)
        if len(matches) > 1:
            descriptions: list[str] = []
            for match in matches:
                row = match.row
                if match.number_range is not None:
                    start, end = match.number_range
                    descriptions.append(
                        f"row {row.spreadsheet_row} value {row.file_name!r} covers "
                        f"{start}\N{EN DASH}{end}"
                    )
                elif isinstance(key, int):
                    descriptions.append(
                        f"row {row.spreadsheet_row} value {row.file_name!r} resolves "
                        f"to {key}"
                    )
                else:
                    descriptions.append(
                        f"row {row.spreadsheet_row} value {row.file_name!r} matches "
                        "directly"
                    )
            lookup = (
                f"file number {key}" if isinstance(key, int) else f"file name {key!r}"
            )
            raise ValueError(
                f"{self.source_name} {lookup} is ambiguous: " + "; ".join(descriptions)
            )
        row = matches[0].row.values

        values_by_header = dict(zip(headers, row, strict=True))
        coordinate_values = {
            destination: values_by_header[resolved_columns[source]]
            for source, destination in self.coordinate_mapping.items()
            if not _is_blank(values_by_header[resolved_columns[source]])
        }
        attribute_values = {
            destination: values_by_header[resolved_columns[source]]
            for source, destination in self.attribute_mapping.items()
            if not _is_blank(values_by_header[resolved_columns[source]])
        }
        return _SpreadsheetMetadataValues(
            coordinate_values,
            attribute_values,
            overwrite=self.overwrite,
        )

    def _normalize_headers(self, raw_headers: Iterable[Any]) -> tuple[str, ...]:
        raw_headers = list(raw_headers)
        while raw_headers and _is_blank(raw_headers[-1]):
            raw_headers.pop()
        if not raw_headers:
            raise ValueError(f"{self.source_name} has no header row")

        headers = tuple(
            _normalize_column_name(value, description="spreadsheet header")
            for value in raw_headers
        )
        if len(set(headers)) != len(headers):
            raise ValueError(f"{self.source_name} has duplicate column headers")
        return headers

    def _resolve_column_name(
        self, requested: str, headers: tuple[str, ...]
    ) -> str | None:
        if requested in headers:
            return requested

        matches = [
            header for header in headers if header.replace("\n", " ") == requested
        ]
        if len(matches) > 1:
            names = ", ".join(repr(value) for value in matches)
            raise ValueError(
                f"{self.source_name} column {requested!r} is ambiguous after replacing "
                f"line breaks with spaces: {names}"
            )
        if matches:
            return matches[0]
        return None

    def _resolve_mapping_columns(self, headers: tuple[str, ...]) -> dict[str, str]:
        requested_columns = {*self.coordinate_mapping, *self.attribute_mapping}
        resolved_columns: dict[str, str] = {}
        missing_columns: set[str] = set()
        for requested in requested_columns:
            resolved = self._resolve_column_name(requested, headers)
            if resolved is None:
                missing_columns.add(requested)
            else:
                resolved_columns[requested] = resolved
        if missing_columns:
            names = ", ".join(repr(value) for value in sorted(missing_columns))
            raise ValueError(f"{self.source_name} is missing columns: {names}")
        return resolved_columns

    def _validate_mapping_columns(self, headers: tuple[str, ...]) -> None:
        self._resolve_mapping_columns(headers)

    def _resolve_file_name_column(self, headers: tuple[str, ...]) -> str:
        if self.file_name_column is None:
            raise ValueError(
                "file_name_column must be provided to apply spreadsheet metadata"
            )
        resolved = self._resolve_column_name(self.file_name_column, headers)
        if resolved is None:
            raise ValueError(
                f"{self.source_name} is missing columns: {self.file_name_column!r}"
            )
        return resolved

    def _build_index(
        self, rows: Iterable[Iterable[Any]]
    ) -> tuple[tuple[str, ...], _SpreadsheetIndex]:
        row_iterator = iter(rows)
        try:
            raw_headers = list(next(row_iterator))
        except StopIteration as exc:
            raise ValueError(f"{self.source_name} is empty") from exc

        headers = self._normalize_headers(raw_headers)

        resolved_file_name_column = self._resolve_file_name_column(headers)
        file_name_index = headers.index(resolved_file_name_column)
        file_names: dict[str, list[_SpreadsheetRow]] = {}
        ranges: list[_FileNumberRange] = []
        first_data_row = 2 if self.row_range is None else self.row_range[0]
        for spreadsheet_row, raw_row in enumerate(row_iterator, start=first_data_row):
            row = list(raw_row)
            if len(row) > len(headers):
                if any(not _is_blank(value) for value in row[len(headers) :]):
                    raise ValueError(
                        f"{self.source_name} row {spreadsheet_row} has values without "
                        "column headers"
                    )
                del row[len(headers) :]
            row.extend([None] * (len(headers) - len(row)))

            if _is_blank(row[file_name_index]):
                continue
            raw_file_name = row[file_name_index]
            file_name = str(raw_file_name).strip()
            if not file_name:
                continue
            row_record = _SpreadsheetRow(spreadsheet_row, raw_file_name, tuple(row))
            try:
                number_range = _normalize_file_number_range(
                    raw_file_name, spreadsheet_row=spreadsheet_row
                )
            except ValueError:
                number_range = None
            if number_range is not None:
                ranges.append(_FileNumberRange(*number_range, row_record))
                continue
            file_names.setdefault(file_name, []).append(row_record)

        return headers, _SpreadsheetIndex(file_names, tuple(ranges))


class ExcelMetadataSource(SpreadsheetMetadataSource):
    """Metadata source backed by a local Excel worksheet.

    Parameters
    ----------
    path
        Path to an ``.xlsx`` or ``.xlsm`` workbook.
    sheet_name
        Worksheet name or zero-based worksheet index.

    Other Parameters
    ----------------
    file_name_column, coordinate_mapping, attribute_mapping, overwrite, row_range
        See :class:`SpreadsheetMetadataSource`.

    Notes
    -----
    Worksheet names and headers are cached, while row content is read from the workbook
    for every metadata lookup. Call :meth:`refresh` to revalidate the selected worksheet
    and header.

    .. versionadded:: 3.25.0
    """

    _cache_row_index = False

    def __init__(
        self,
        path: str | pathlib.Path,
        *,
        sheet_name: str | int = 0,
        file_name_column: str | None = None,
        coordinate_mapping: Mapping[str, str] | None = None,
        attribute_mapping: Mapping[str, str] | None = None,
        overwrite: bool = False,
        row_range: tuple[int, int] | None = None,
    ) -> None:
        self.path = pathlib.Path(path)
        self._selected_sheet_name: str | None = None
        if self.path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError("Excel metadata files must use .xlsx or .xlsm")
        super().__init__(
            file_name_column=file_name_column,
            coordinate_mapping=coordinate_mapping,
            attribute_mapping=attribute_mapping,
            sheet_name=sheet_name,
            overwrite=overwrite,
            row_range=row_range,
        )

    @property
    def source_name(self) -> str:
        return f"Excel metadata source {self.path!s}"

    def _load_workbook(self):
        if not importlib.util.find_spec("openpyxl"):
            raise ImportError("`openpyxl` needs to be installed to read Excel metadata")
        import openpyxl

        return openpyxl.load_workbook(
            self.path, read_only=True, data_only=True, keep_links=False
        )

    def _select_worksheet(self, workbook):
        if self._selected_sheet_name is not None:
            try:
                return workbook[self._selected_sheet_name]
            except KeyError as exc:
                raise ValueError(
                    f"Worksheet {self._selected_sheet_name!r} was not found in "
                    f"{self.path!s}"
                ) from exc

        sheet_name = typing.cast("str | int", self.sheet_name)
        try:
            if isinstance(sheet_name, int):
                worksheet = workbook.worksheets[sheet_name]
            else:
                worksheet = workbook[sheet_name]
        except (IndexError, KeyError) as exc:
            raise ValueError(
                f"Worksheet {sheet_name!r} was not found in {self.path!s}"
            ) from exc
        self._selected_sheet_name = worksheet.title
        return worksheet

    def _clear_source_cache(self) -> None:
        self._selected_sheet_name = None

    def _read_sheet_names(self) -> list[str]:
        workbook = self._load_workbook()
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    def _read_header(self) -> list[Any]:
        workbook = self._load_workbook()
        try:
            worksheet = self._select_worksheet(workbook)
            row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            return list(row)
        finally:
            workbook.close()

    def _read_rows(self) -> list[list[Any]]:
        workbook = self._load_workbook()
        try:
            worksheet = self._select_worksheet(workbook)
            if self._headers is None:
                header = list(
                    next(
                        worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
                    )
                )
            else:
                header = list(self._headers)
            min_row = 2 if self.row_range is None else self.row_range[0]
            max_row = None if self.row_range is None else self.row_range[1]
            data_rows = [
                list(row)
                for row in worksheet.iter_rows(
                    min_row=min_row, max_row=max_row, values_only=True
                )
            ]
            return [header, *data_rows]
        finally:
            workbook.close()

    def refresh(self) -> None:
        """Refresh the cached worksheet selection and column names."""
        with self._lock:
            previous_selected_sheet = self._selected_sheet_name
            try:
                super().refresh()
            except Exception:
                self._selected_sheet_name = previous_selected_sheet
                raise
