from __future__ import annotations

import concurrent.futures
import json
import pathlib
import threading
import typing

import numpy as np
import pytest
import requests
import xarray as xr

import erlab
from erlab.io.dataloader import LoaderBase
from erlab.io.metadata import (
    ExcelMetadataSource,
    GoogleSheetsMetadataSource,
    SpreadsheetMetadataSource,
)
from erlab.io.metadata import _core as metadata_core
from erlab.io.metadata import _google as google_impl


class _RowsMetadataSource(SpreadsheetMetadataSource):
    def __init__(self, rows: list[list[typing.Any]], **kwargs) -> None:
        self.rows = rows
        self.read_count = 0
        super().__init__(**kwargs)

    @property
    def source_name(self) -> str:
        return "test spreadsheet"

    def _read_rows(self) -> list[list[typing.Any]]:
        self.read_count += 1
        rows = [row.copy() for row in self.rows]
        if self.row_range is None:
            return rows
        start, end = self.row_range
        return [rows[0], *rows[start - 1 : end]]

    def _read_sheet_names(self) -> list[str]:
        return ["Sheet1"]


class _StrictMetadataLoader(LoaderBase):
    name = "_strict_metadata"
    extensions: typing.ClassVar[set[str]] = {".nc"}
    strict_validation = True

    def identify(self, num: int, data_dir: str | pathlib.Path):
        return [pathlib.Path(data_dir) / f"scan_{num}.nc"], {}

    def load_single(self, file_path: str | pathlib.Path) -> xr.DataArray:
        return xr.DataArray(
            np.arange(2.0),
            dims="eV",
            coords={"eV": [-1.0, 0.0], "beta": 0.0, "delta": 0.0, "xi": 0.0},
        )


class _InferringMetadataLoader(_StrictMetadataLoader):
    name = "_inferring_metadata"

    def infer_index(self, name: str) -> tuple[int | None, dict[str, typing.Any]]:
        prefix, separator, number = name.rpartition("_")
        if prefix == "scan" and separator and number.isdigit():
            return int(number), {}
        return None, {}


class _FakeResponse:
    def __init__(
        self,
        text: str = "",
        *,
        status_code: int = 200,
    ) -> None:
        self.text = text
        self.status_code = status_code

    def raise_for_status(self) -> None:
        if self.status_code >= 400:
            raise requests.HTTPError("request failed")


def _google_sheet_list(*properties: tuple[str, int]) -> str:
    return "".join(
        "items.push({name: "
        f'{json.dumps(name)}, pageUrl: "", gid: "{gid}", '
        "initialSheet: false});"
        for name, gid in properties
    )


def _source(
    rows: list[list[typing.Any]], *, overwrite: bool = False
) -> _RowsMetadataSource:
    return _RowsMetadataSource(
        rows,
        file_name_column="File",
        coordinate_mapping={"Temperature": "sample_temp", "Energy": "hv"},
        attribute_mapping={"Mode": "mode"},
        overwrite=overwrite,
    )


def test_source_snapshot_and_refresh() -> None:
    source = _source(
        [
            ["File", "Temperature", "Energy", "Mode"],
            ["0042", 20.0, 21.2, "mapping"],
            [43.0, 30.0, 40.8, "cut"],
        ]
    )

    values = source._metadata_for_file_number(42)
    assert values is not None
    assert values.coordinate_values == {"sample_temp": 20.0, "hv": 21.2}
    assert values.attribute_values == {"mode": "mapping"}
    assert source.read_count == 1

    source.rows[1][1] = 25.0
    assert source._metadata_for_file_number(42).coordinate_values["sample_temp"] == 20.0
    assert source.read_count == 1

    source.refresh()
    assert source._metadata_for_file_number(42).coordinate_values["sample_temp"] == 25.0
    assert source.read_count == 2


def test_failed_refresh_preserves_cached_snapshot() -> None:
    source = _source(
        [["File", "Temperature", "Energy", "Mode"], [1, 20.0, 21.2, "cut"]]
    )
    values = source._metadata_for_file_number(1)
    assert values is not None

    source.rows.append([2, 30.0, 40.8, "map", "unexpected"])
    with pytest.raises(ValueError, match="row 3 has values without column headers"):
        source.refresh()

    cached_values = source._metadata_for_file_number(1)
    assert cached_values is not None
    assert cached_values.coordinate_values["sample_temp"] == 20.0


def test_source_column_names_use_header_cache() -> None:
    source = _source(
        [
            ["File", "Temperature", "Energy", "Mode"],
            [1, 20.0, 21.2, "mapping"],
        ]
    )

    column_names = source.get_column_names()
    assert column_names == ["File", "Temperature", "Energy", "Mode"]
    assert source.read_count == 1

    column_names.pop()
    assert source.get_column_names() == ["File", "Temperature", "Energy", "Mode"]
    assert source._metadata_for_file_number(1) is not None
    assert source.read_count == 2

    source.rows[0].append("Comment")
    assert source.get_column_names() == ["File", "Temperature", "Energy", "Mode"]
    source.refresh()
    assert source.get_column_names() == [
        "File",
        "Temperature",
        "Energy",
        "Mode",
        "Comment",
    ]
    assert source.read_count == 3


def test_source_discovers_headers_without_indexing_rows() -> None:
    source = _RowsMetadataSource(
        [["File", "Temperature"], ["f_0001", 20.0], ["f_0001", 30.0]],
        file_name_column="File",
        coordinate_mapping={"Temperature": "sample_temp"},
    )

    assert source.get_column_names() == ["File", "Temperature"]
    with pytest.raises(ValueError, match=r"rows 2 and 3.*both match file name"):
        source._metadata_for_file_name(
            "f_0001", lambda: pytest.fail("numeric inference must not run")
        )


def test_source_can_be_inspected_without_file_name_column() -> None:
    source = _RowsMetadataSource([["File", "Temperature"], [1, 20.0]])

    assert source.get_sheet_names() == ["Sheet1"]
    assert source.get_column_names() == ["File", "Temperature"]
    with pytest.raises(ValueError, match="coordinate or attribute mapping"):
        source._metadata_for_file_number(1)


def test_source_mappings_are_immutable() -> None:
    source = _source([["File", "Temperature", "Energy", "Mode"]])

    with pytest.raises(TypeError):
        source.coordinate_mapping["Missing"] = "missing"  # type: ignore[index]
    with pytest.raises(AttributeError):
        source.coordinate_mapping = {}  # type: ignore[misc]


def test_source_matches_line_breaks_in_column_names() -> None:
    source = _RowsMetadataSource(
        [
            ["File\nName", "Photon\nEnergy", "Acquisition\nMode"],
            ["f_0007.pxt", 21.2, "mapping"],
        ],
        file_name_column="File Name",
        coordinate_mapping={"Photon Energy": "hv"},
        attribute_mapping={"Acquisition Mode": "mode"},
    )

    assert source.get_column_names() == [
        "File\nName",
        "Photon\nEnergy",
        "Acquisition\nMode",
    ]
    values = source._metadata_for_file_number(7)
    assert values is not None
    assert values.coordinate_values == {"hv": 21.2}
    assert values.attribute_values == {"mode": "mapping"}


def test_source_prefers_exact_column_name_over_line_break_alias() -> None:
    source = _RowsMetadataSource(
        [
            ["File Name", "File\nName", "Photon Energy", "Photon\nEnergy"],
            [7, 8, 21.2, 40.8],
        ],
        file_name_column="File Name",
        coordinate_mapping={"Photon Energy": "hv"},
    )

    values = source._metadata_for_file_number(7)
    assert values is not None
    assert values.coordinate_values == {"hv": 21.2}
    assert source._metadata_for_file_number(8) is None


@pytest.mark.parametrize(
    ("rows", "match"),
    [
        ([], "is empty"),
        ([[None, ""]], "has no header row"),
        (
            [["File", "Temperature", "Energy", "Mode", "File"]],
            "duplicate column headers",
        ),
        (
            [["File", "Temperature", "Mode"], [1, 20, "a"]],
            "missing columns: 'Energy'",
        ),
        (
            [
                ["File", "Temperature", "Energy", "Mode", None],
                [1, 20, 21, "a", "extra"],
            ],
            "values without column headers",
        ),
    ],
)
def test_source_rejects_invalid_tables(
    rows: list[list[typing.Any]], match: str
) -> None:
    with pytest.raises(ValueError, match=match):
        _source(rows).refresh()


def test_source_skips_rows_without_file_numbers() -> None:
    source = _source(
        [
            ["File", "Temperature", "Energy", "Mode"],
            [None, 20.0, 21.2, "comment row"],
            [],
            [1, 30.0, 40.8, "cut"],
        ]
    )

    values = source._metadata_for_file_number(1)
    assert values is not None
    assert values.coordinate_values == {"sample_temp": 30.0, "hv": 40.8}
    assert values.attribute_values == {"mode": "cut"}


@pytest.mark.parametrize(
    ("spreadsheet_value", "file_number"),
    [
        ("f_0001", 1),
        ("scan2026_run_0042.pxt", 42),
        ("/data/scan_0003.nc", 3),
        (r"C:\data\scan_0004.h5", 4),
        ("0005", 5),
        (6.0, 6),
    ],
)
def test_source_parses_simple_filenames(
    spreadsheet_value: object, file_number: int
) -> None:
    source = _source(
        [
            ["File", "Temperature", "Energy", "Mode"],
            [spreadsheet_value, 30.0, 40.8, "cut"],
        ]
    )

    values = source._metadata_for_file_number(file_number)
    assert values is not None
    assert values.coordinate_values == {"sample_temp": 30.0, "hv": 40.8}


def test_source_prefers_trimmed_file_name_without_number_parsing(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _source(
        [
            ["File", "Temperature", "Energy", "Mode"],
            ["  sample-final  ", 20.0, 21.2, "a"],
        ]
    )
    monkeypatch.setattr(
        metadata_core,
        "_normalize_file_number",
        lambda *_args, **_kwargs: pytest.fail("number parsing must not run"),
    )

    values, lookup = source._metadata_for_file_name(
        "sample-final", lambda: pytest.fail("numeric inference must not run")
    )

    assert lookup == "sample-final"
    assert values is not None
    assert values.coordinate_values == {"sample_temp": 20.0, "hv": 21.2}


def test_source_reports_ambiguous_numeric_fallback() -> None:
    source = _source(
        [
            ["File", "Temperature", "Energy", "Mode"],
            ["sample_a_0001", 20.0, 21.2, "a"],
            ["sample_b_0001", 30.0, 40.8, "b"],
        ]
    )

    with pytest.raises(
        ValueError,
        match=r"rows 2 and 3.*both resolve to 1",
    ):
        source._metadata_for_file_number(1)


def test_source_skips_unparseable_file_number_rows() -> None:
    source = _source(
        [
            ["File", "Temperature", "Energy", "Mode"],
            [1.5, 10.0, 21.2, "invalid number"],
            ["notes", 20.0, 21.2, "text"],
            ["f-0002", 25.0, 21.2, "unsupported name"],
            ["f_0003.pxt", 30.0, 40.8, "cut"],
        ]
    )

    values = source._metadata_for_file_number(3)
    assert values is not None
    assert values.coordinate_values == {"sample_temp": 30.0, "hv": 40.8}
    assert source._metadata_for_file_number(2) is None


def test_source_trims_blank_cells_without_headers() -> None:
    source = _source(
        [
            ["File", "Temperature", "Energy", "Mode", None],
            [1, 30.0, 40.8, "cut", ""],
        ]
    )

    values = source._metadata_for_file_number(1)
    assert values is not None
    assert values.coordinate_values == {"sample_temp": 30.0, "hv": 40.8}


def test_source_rejects_invalid_mappings() -> None:
    source_without_mappings = _RowsMetadataSource(
        [["File", "Temperature"]], file_name_column="File"
    )
    assert source_without_mappings.get_column_names() == ["File", "Temperature"]
    with pytest.raises(ValueError, match="At least one"):
        source_without_mappings._metadata_for_file_number(1)

    with pytest.raises(ValueError, match="overlapping destinations"):
        _RowsMetadataSource(
            [["File", "A", "B"]],
            file_name_column="File",
            coordinate_mapping={"A": "value"},
            attribute_mapping={"B": "value"},
        )

    with pytest.raises(ValueError, match="Duplicate coordinate mapping"):
        _RowsMetadataSource(
            [["File", "A"]],
            file_name_column="File",
            coordinate_mapping={" A": "first", "A": "second"},
        )


@pytest.mark.parametrize(
    ("kwargs", "error"),
    [
        ({"file_name_column": ""}, ValueError),
        ({"coordinate_mapping": {"Temperature": ""}}, ValueError),
        ({"overwrite": 1}, TypeError),
        ({"row_range": [2, 3]}, TypeError),
    ],
)
def test_source_rejects_invalid_options(
    kwargs: dict[str, typing.Any], error: type[Exception]
) -> None:
    options: dict[str, typing.Any] = {
        "file_name_column": "File",
        "coordinate_mapping": {"Temperature": "sample_temp"},
    }
    options.update(kwargs)
    with pytest.raises(error):
        _RowsMetadataSource([["File", "Temperature"]], **options)


@pytest.mark.parametrize(
    ("sheet_name", "error"),
    [("", ValueError), (-1, ValueError), (True, TypeError)],
)
def test_source_rejects_invalid_sheet_name(
    sheet_name: typing.Any, error: type[Exception]
) -> None:
    with pytest.raises(error, match="sheet_name"):
        _RowsMetadataSource(
            [["File", "Temperature"]],
            file_name_column="File",
            coordinate_mapping={"Temperature": "sample_temp"},
            sheet_name=sheet_name,
        )


def test_source_limits_duplicate_check_to_row_range() -> None:
    source = _RowsMetadataSource(
        [
            ["File", "Temperature"],
            [1, 20.0],
            [2, 25.0],
            [1, 30.0],
        ],
        file_name_column="File",
        coordinate_mapping={"Temperature": "sample_temp"},
        row_range=(4, 4),
    )

    values = source._metadata_for_file_number(1)
    assert values is not None
    assert values.coordinate_values == {"sample_temp": 30.0}


@pytest.mark.parametrize(
    ("file_number", "reason"),
    [
        (True, "boolean values are not valid"),
        (-1, "negative file numbers are not allowed"),
        (1.5, "numeric values must be integers"),
        (float("nan"), "numeric values must be finite"),
        ("  ", "empty after trimming whitespace"),
        ("not_a_file", "expected decimal digits or a simple filename"),
        (object(), "unsupported value type 'object'"),
    ],
)
def test_source_rejects_invalid_lookup_file_number(
    file_number: object, reason: str
) -> None:
    with pytest.raises(ValueError, match=reason) as exc_info:
        _source([["File", "Temperature", "Energy", "Mode"]])._metadata_for_file_number(
            file_number
        )
    assert repr(file_number) in str(exc_info.value)


def test_file_number_parse_error_includes_spreadsheet_row() -> None:
    with pytest.raises(ValueError, match="spreadsheet row 42") as exc_info:
        metadata_core._normalize_file_number("invalid-name", spreadsheet_row=42)

    assert str(exc_info.value) == (
        "File-name cell 'invalid-name' in spreadsheet row 42 was rejected: expected "
        "decimal digits or a simple filename ending in digits, such as 'f_0001.pxt'"
    )


def test_file_number_parse_error_for_excessively_long_number() -> None:
    with pytest.raises(ValueError, match="numeric portion is too long"):
        metadata_core._normalize_file_number("1" * 10_000, spreadsheet_row=2)


@pytest.mark.parametrize("row_range", [(1, 2), (4, 3)])
def test_source_rejects_invalid_row_range(row_range: tuple[int, int]) -> None:
    with pytest.raises(ValueError, match="row_range"):
        _RowsMetadataSource(
            [["File", "Temperature"]],
            file_name_column="File",
            coordinate_mapping={"Temperature": "sample_temp"},
            row_range=row_range,
        )


def test_values_apply_without_overwrite() -> None:
    values = _source(
        [["File", "Temperature", "Energy", "Mode"], [1, 20.0, 21.2, "sheet"]]
    )._metadata_for_file_number(1)
    assert values is not None

    data = xr.DataArray(
        np.ones(2),
        dims="x",
        coords={"x": [0, 1], "hv": np.nan},
        attrs={"mode": "existing"},
    )
    result = values.apply(data)

    assert result["sample_temp"].item() == 20.0
    assert np.isnan(result["hv"].item())
    assert result.attrs["mode"] == "existing"


def test_values_apply_with_overwrite_and_structures() -> None:
    values = _source(
        [["File", "Temperature", "Energy", "Mode"], [1, 20.0, 21.2, "sheet"]],
        overwrite=True,
    )._metadata_for_file_number(1)
    assert values is not None

    dataset = xr.Dataset(
        {"a": ("x", [1.0, 2.0]), "b": ("x", [3.0, 4.0])},
        coords={"x": [0, 1], "hv": 50.0},
        attrs={"dataset_attr": "preserved"},
    )
    result = values.apply(dataset)
    assert result.attrs == {"dataset_attr": "preserved"}
    for data in result.data_vars.values():
        assert data["sample_temp"].item() == 20.0
        assert data["hv"].item() == 21.2
        assert data.attrs["mode"] == "sheet"

    tree = xr.DataTree.from_dict({"/group": dataset})
    tree_result = values.apply(tree)
    assert tree_result["/group"]["a"]["hv"].item() == 21.2


def test_values_parse_numeric_and_boolean_strings() -> None:
    source = _RowsMetadataSource(
        [
            ["File", "Energy", "Enabled", "Label", "Configuration", "Active"],
            [1, " 21.2 ", "TRUE", "001", "2", "FALSE"],
        ],
        file_name_column="File",
        coordinate_mapping={"Energy": "hv", "Enabled": "enabled"},
        attribute_mapping={
            "Label": "label",
            "Configuration": "configuration",
            "Active": "active",
        },
        overwrite=True,
    )
    values = source._metadata_for_file_number(1)
    assert values is not None

    result = values.apply(
        xr.DataArray([1.0], dims="x", attrs={"configuration": 1, "active": True})
    )
    assert result["hv"].item() == 21.2
    assert result["enabled"].item() is True
    assert result.attrs["label"] == "001"
    assert result.attrs["configuration"] == 2
    assert result.attrs["active"] is False


def test_values_warn_and_skip_invalid_numeric_replacements() -> None:
    source = _RowsMetadataSource(
        [
            ["File", "Temperature", "Energy", "Configuration", "Count"],
            [1, "unknown", "not recorded", "invalid", "TRUE"],
        ],
        file_name_column="File",
        coordinate_mapping={"Temperature": "sample_temp", "Energy": "hv"},
        attribute_mapping={"Configuration": "configuration", "Count": "count"},
        overwrite=True,
    )
    values = source._metadata_for_file_number(1)
    assert values is not None
    data = xr.DataArray(
        [1.0],
        dims="x",
        coords={"sample_temp": 20.0, "hv": 21.2},
        attrs={"configuration": 1, "count": 2},
    )

    with pytest.warns(UserWarning, match="Could not convert") as warnings:
        result = values.apply(data)

    assert len(warnings) == 4
    xr.testing.assert_identical(result, data)


def test_values_warn_and_skip_coordinate_application_error() -> None:
    source = _RowsMetadataSource(
        [["File", "Invalid"], [1, {"unsupported": "coordinate value"}]],
        file_name_column="File",
        coordinate_mapping={"Invalid": "invalid"},
    )
    values = source._metadata_for_file_number(1)
    assert values is not None
    data = xr.DataArray([1.0], dims="x")

    with pytest.warns(UserWarning, match="Could not apply"):
        result = values.apply(data)

    xr.testing.assert_identical(result, data)


def test_values_reject_dimension_coordinate_overwrite() -> None:
    source = _RowsMetadataSource(
        [["File", "Position"], [1, 2.0]],
        file_name_column="File",
        coordinate_mapping={"Position": "x"},
        overwrite=True,
    )
    values = source._metadata_for_file_number(1)
    assert values is not None

    with pytest.raises(ValueError, match="dimension coordinate 'x'"):
        values.apply(xr.DataArray(np.ones(2), dims="x", coords={"x": [0, 1]}))


def test_values_preserve_dimension_and_reject_non_scalar_coordinate() -> None:
    source = _RowsMetadataSource(
        [["File", "Position"], [1, 2.0]],
        file_name_column="File",
        coordinate_mapping={"Position": "x"},
    )
    values = source._metadata_for_file_number(1)
    assert values is not None
    data = xr.DataArray(np.ones(2), dims="x", coords={"x": [0, 1]})
    xr.testing.assert_identical(values.apply(data), data)

    overwrite_source = _RowsMetadataSource(
        [["File", "Position"], [1, 2.0]],
        file_name_column="File",
        coordinate_mapping={"Position": "position"},
        overwrite=True,
    )
    values = overwrite_source._metadata_for_file_number(1)
    assert values is not None
    non_scalar = xr.DataArray(
        np.ones(2), dims="x", coords={"position": ("x", [0.0, 1.0])}
    )
    with pytest.raises(ValueError, match="non-scalar coordinate 'position'"):
        values.apply(non_scalar)


def test_excel_metadata_source(tmp_path: pathlib.Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.active.title = "Other"
    worksheet = workbook.create_sheet("Metadata")
    worksheet.append(["File", "Temperature", "Energy", "Mode"])
    worksheet.append([7, 35.0, 40.8, "cut"])
    path = tmp_path / "metadata.xlsx"
    workbook.save(path)
    workbook.close()

    source = ExcelMetadataSource(
        path,
        sheet_name="Metadata",
        file_name_column="File",
        coordinate_mapping={"Temperature": "sample_temp", "Energy": "hv"},
        attribute_mapping={"Mode": "mode"},
    )
    assert source.get_sheet_names() == ["Other", "Metadata"]
    assert source.get_column_names() == ["File", "Temperature", "Energy", "Mode"]
    values = source._metadata_for_file_number(7)
    assert values is not None
    assert source.source_name == f"Excel metadata source {path!s}"
    assert values.coordinate_values == {"sample_temp": 35.0, "hv": 40.8}
    assert values.attribute_values == {"mode": "cut"}

    by_index = ExcelMetadataSource(
        path,
        sheet_name=1,
        file_name_column="File",
        coordinate_mapping={"Temperature": "sample_temp"},
    )
    assert by_index._metadata_for_file_number(7) is not None


def test_excel_source_errors(tmp_path: pathlib.Path) -> None:
    with pytest.raises(ValueError, match=r"\.xlsx or \.xlsm"):
        ExcelMetadataSource(
            tmp_path / "metadata.xls",
            file_name_column="File",
            coordinate_mapping={"Temperature": "sample_temp"},
        )

    import openpyxl

    path = tmp_path / "metadata.xlsx"
    workbook = openpyxl.Workbook()
    workbook.save(path)
    workbook.close()
    source = ExcelMetadataSource(
        path,
        sheet_name="Missing",
        file_name_column="File",
        coordinate_mapping={"Temperature": "sample_temp"},
    )
    with pytest.raises(ValueError, match="Worksheet 'Missing' was not found"):
        source.refresh()


def test_excel_optional_dependency_error(
    tmp_path: pathlib.Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = ExcelMetadataSource(
        tmp_path / "metadata.xlsx",
        file_name_column="File",
        coordinate_mapping={"Temperature": "sample_temp"},
    )
    monkeypatch.setattr(metadata_core.importlib.util, "find_spec", lambda _name: None)

    with pytest.raises(ImportError, match="`openpyxl` needs to be installed"):
        source.refresh()


def test_loader_applies_metadata_before_strict_validation(
    tmp_path: pathlib.Path,
) -> None:
    (tmp_path / "scan_1.nc").touch()
    source = _RowsMetadataSource(
        [
            ["File", "Energy", "Temperature", "Configuration"],
            ["scan_1", 21.2, 30.0, 1],
        ],
        file_name_column="File",
        coordinate_mapping={"Energy": "hv", "Temperature": "sample_temp"},
        attribute_mapping={"Configuration": "configuration"},
    )
    loader = _StrictMetadataLoader()

    result = loader.load(1, tmp_path, metadata=source)
    assert result["hv"].item() == 21.2
    assert result["sample_temp"].item() == 30.0
    assert result.attrs["configuration"] == 1

    path_result = loader.load(tmp_path / "scan_1.nc", metadata=source)
    xr.testing.assert_identical(path_result, result)


def test_loader_infers_spreadsheet_file_number_from_path(
    tmp_path: pathlib.Path,
) -> None:
    path = tmp_path / "scan_1.nc"
    path.touch()
    source = _RowsMetadataSource(
        [["File", "Energy", "Temperature", "Configuration"], [1, 21.2, 30.0, 1]],
        file_name_column="File",
        coordinate_mapping={"Energy": "hv", "Temperature": "sample_temp"},
        attribute_mapping={"Configuration": "configuration"},
    )

    result = _InferringMetadataLoader().load(path, metadata=source)

    assert result["hv"].item() == 21.2
    assert result["sample_temp"].item() == 30.0
    assert result.attrs["configuration"] == 1


def test_concurrent_loads_keep_metadata_isolated(
    tmp_path: pathlib.Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path_a = tmp_path / "a.nc"
    path_b = tmp_path / "b.nc"
    path_a.touch()
    path_b.touch()
    entered_a = threading.Event()
    entered_b = threading.Event()
    finished_a = threading.Event()
    loader = _StrictMetadataLoader()
    loader.skip_validate = True

    def load_single(file_path: str | pathlib.Path) -> xr.DataArray:
        if pathlib.Path(file_path).stem == "a":
            entered_a.set()
            if not entered_b.wait(timeout=5):
                raise TimeoutError("the second load did not start")
        else:
            if not entered_a.wait(timeout=5):
                raise TimeoutError("the first load did not start")
            entered_b.set()
            if not finished_a.wait(timeout=5):
                raise TimeoutError("the first load did not finish")
        return xr.DataArray([1.0], dims="x")

    monkeypatch.setattr(loader, "load_single", load_single)
    source_a = _RowsMetadataSource(
        [["File", "Label"], [1, "A"]],
        file_name_column="File",
        attribute_mapping={"Label": "label"},
    )
    source_b = _RowsMetadataSource(
        [["File", "Label"], [2, "B"]],
        file_name_column="File",
        attribute_mapping={"Label": "label"},
    )

    def load_a() -> xr.DataArray:
        result = loader.load(path_a, metadata=source_a, file_number=1)
        finished_a.set()
        return typing.cast("xr.DataArray", result)

    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        future_a = executor.submit(load_a)
        future_b = executor.submit(
            loader.load, path_b, metadata=source_b, file_number=2
        )
        result_a = future_a.result(timeout=10)
        result_b = typing.cast("xr.DataArray", future_b.result(timeout=10))

    assert result_a.attrs["label"] == "A"
    assert result_b.attrs["label"] == "B"


def test_loader_metadata_arguments(tmp_path: pathlib.Path) -> None:
    (tmp_path / "scan_1.nc").touch()
    source = _RowsMetadataSource(
        [["File", "Energy"], [1, 21.2]],
        file_name_column="File",
        coordinate_mapping={"Energy": "hv"},
    )
    loader = _StrictMetadataLoader()

    with pytest.raises(ValueError, match="file_number requires"):
        loader.load(tmp_path / "scan_1.nc", file_number=1)
    with pytest.raises(
        ValueError,
        match=r"file_number is required.*_strict_metadata.*could not infer",
    ):
        loader.load(tmp_path / "scan_1.nc", metadata=source)
    with pytest.raises(ValueError, match="must be omitted"):
        loader.load(1, tmp_path, metadata=source, file_number=1)
    with pytest.raises(TypeError, match="SpreadsheetMetadataSource"):
        loader.load(1, tmp_path, metadata=object())


def test_top_level_load_and_combine_false_metadata(
    example_loader,
    example_data_dir: pathlib.Path,
) -> None:
    source = _RowsMetadataSource(
        [["File", "Label", "Extra"], [1, "scan one", 7.0]],
        file_name_column="File",
        coordinate_mapping={"Extra": "spreadsheet_extra"},
        attribute_mapping={"Label": "spreadsheet_label"},
    )

    with erlab.io.loader_context("example", example_data_dir):
        results = erlab.io.load(1, metadata=source, combine=False, progress=False)

    assert isinstance(results, list)
    assert len(results) > 1
    for result in results:
        assert result["spreadsheet_extra"].item() == 7.0
        assert result.attrs["spreadsheet_label"] == "scan one"


def test_loader_warns_for_missing_spreadsheet_row(
    example_loader,
    example_data_dir: pathlib.Path,
) -> None:
    source = _RowsMetadataSource(
        [["File", "Extra"], [99, 7.0]],
        file_name_column="File",
        coordinate_mapping={"Extra": "spreadsheet_extra"},
    )

    with (
        erlab.io.loader_context("example", example_data_dir),
        pytest.warns(UserWarning, match="No spreadsheet metadata found"),
    ):
        result = erlab.io.load(2, metadata=source)
    assert "spreadsheet_extra" not in result.coords


@pytest.mark.parametrize(
    "share_url",
    [
        "",
        "http://docs.google.com/spreadsheets/d/abc/edit",
        "https://example.com/spreadsheets/d/abc/edit",
        "https://docs.google.com/document/d/abc/edit",
        "https://docs.google.com/spreadsheets/d//edit",
        "https://docs.google.com/spreadsheets/d/abc/edit#gid=invalid",
        "https://docs.google.com/spreadsheets/d/abc/edit#gid=-1",
    ],
)
def test_google_source_rejects_invalid_share_links(share_url: str) -> None:
    with pytest.raises(ValueError, match="share_url"):
        GoogleSheetsMetadataSource(
            share_url,
            file_name_column="File",
            coordinate_mapping={"Energy": "hv"},
        )


@pytest.mark.parametrize(
    ("kwargs", "error"),
    [
        ({"sheet_name": 1}, TypeError),
        ({"timeout": True}, TypeError),
        ({"timeout": 0}, ValueError),
    ],
)
def test_google_source_rejects_invalid_options(
    kwargs: dict[str, typing.Any], error: type[Exception]
) -> None:
    options: dict[str, typing.Any] = {
        "file_name_column": "File",
        "coordinate_mapping": {"Energy": "hv"},
    }
    options.update(kwargs)
    with pytest.raises(error):
        GoogleSheetsMetadataSource(
            "https://docs.google.com/spreadsheets/d/sheet-id/edit", **options
        )


def test_google_optional_dependency_error(monkeypatch: pytest.MonkeyPatch) -> None:
    source = GoogleSheetsMetadataSource(
        "https://docs.google.com/spreadsheets/d/sheet-id/edit",
        sheet_name="Sheet1",
        file_name_column="File",
        coordinate_mapping={"Energy": "hv"},
    )
    monkeypatch.setattr(google_impl.importlib.util, "find_spec", lambda _name: None)

    with pytest.raises(ImportError, match="`requests` needs to be installed"):
        source.refresh()


def test_google_source_reads_linked_tab(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    calls: list[tuple[str, dict[str, typing.Any]]] = []

    def fake_get(url: str, **kwargs) -> _FakeResponse:
        calls.append((url, kwargs))
        if url.endswith("/htmlview"):
            return _FakeResponse(
                _google_sheet_list(("Linked Sheet", 123), ("Other Sheet", 456))
            )
        return _FakeResponse(
            "\ufeffFile,Temperature,Mode,Enabled,Label\n"
            '007,35.0,"cut, high resolution",TRUE,001\n'
        )

    monkeypatch.setattr(requests, "get", fake_get)
    source = GoogleSheetsMetadataSource(
        "https://docs.google.com/spreadsheets/d/sheet-id/edit?usp=sharing#gid=123",
        file_name_column="File",
        coordinate_mapping={"Temperature": "sample_temp"},
        attribute_mapping={
            "Mode": "mode",
            "Enabled": "enabled",
            "Label": "label",
        },
    )

    assert source.get_sheet_names() == ["Linked Sheet", "Other Sheet"]
    assert source.get_column_names() == [
        "File",
        "Temperature",
        "Mode",
        "Enabled",
        "Label",
    ]
    values = source._metadata_for_file_number(7)
    assert values is not None
    assert values.coordinate_values == {"sample_temp": 35.0}
    assert values.attribute_values == {
        "mode": "cut, high resolution",
        "enabled": True,
        "label": "001",
    }
    assert source.spreadsheet_id == "sheet-id"
    assert source.gid == 123
    assert source.source_name == (
        "Google Sheets metadata source "
        "'https://docs.google.com/spreadsheets/d/sheet-id/edit?usp=sharing#gid=123'"
    )
    assert calls == [
        (
            "https://docs.google.com/spreadsheets/d/sheet-id/htmlview",
            {"params": {}, "timeout": 10.0},
        ),
        (
            "https://docs.google.com/spreadsheets/d/sheet-id/gviz/tq",
            {
                "params": {
                    "tqx": "out:csv",
                    "headers": "0",
                    "gid": "123",
                    "range": "1:1",
                },
                "timeout": 10.0,
            },
        ),
        (
            "https://docs.google.com/spreadsheets/d/sheet-id/gviz/tq",
            {
                "params": {"tqx": "out:csv", "headers": "1", "gid": "123"},
                "timeout": 10.0,
            },
        ),
    ]

    source._metadata_for_file_number(7)
    assert len(calls) == 4
    source.refresh()
    assert len(calls) == 6


@pytest.mark.parametrize(
    ("share_url", "expected_gid"),
    [
        (
            "https://docs.google.com/spreadsheets/d/sheet-id/edit?gid=456",
            "456",
        ),
        (
            "https://docs.google.com/spreadsheets/d/sheet-id/edit?gid=456#gid=123",
            "123",
        ),
    ],
)
def test_google_source_reads_gid_from_link(
    monkeypatch: pytest.MonkeyPatch,
    share_url: str,
    expected_gid: str,
) -> None:
    params: list[dict[str, str]] = []

    def fake_get(url: str, **kwargs) -> _FakeResponse:
        if url.endswith("/htmlview"):
            return _FakeResponse(
                _google_sheet_list(("Linked Sheet", int(expected_gid)))
            )
        params.append(kwargs["params"])
        return _FakeResponse("File,Energy\n1,21.2\n")

    monkeypatch.setattr(requests, "get", fake_get)
    source = GoogleSheetsMetadataSource(
        share_url,
        file_name_column="File",
        coordinate_mapping={"Energy": "hv"},
    )

    assert source._metadata_for_file_number(1) is not None
    assert params[0]["gid"] == expected_gid


def test_google_source_selects_worksheet_by_name(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    params: list[dict[str, str]] = []

    def fake_get(url: str, **kwargs) -> _FakeResponse:
        if url.endswith("/htmlview"):
            return _FakeResponse(
                _google_sheet_list(("Linked Sheet", 123), ("Other Sheet", 456))
            )
        params.append(kwargs["params"])
        return _FakeResponse("File,Energy\n1,21.2\n")

    monkeypatch.setattr(requests, "get", fake_get)
    source = GoogleSheetsMetadataSource(
        "https://docs.google.com/spreadsheets/d/sheet-id/edit#gid=123",
        sheet_name="Other Sheet",
        file_name_column="File",
        coordinate_mapping={"Energy": "hv"},
    )

    assert source._metadata_for_file_number(1) is not None
    assert params == [{"tqx": "out:csv", "headers": "1", "sheet": "Other Sheet"}]


def test_google_source_rejects_missing_worksheet_name(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    params: list[dict[str, str]] = []

    def fake_get(url: str, **kwargs) -> _FakeResponse:
        if url.endswith("/htmlview"):
            return _FakeResponse(
                _google_sheet_list(("Linked Sheet", 123), ("Other Sheet", 456))
            )
        params.append(kwargs["params"])
        # Google silently returns the first worksheet for an invalid sheet name.
        return _FakeResponse("File,Energy\n1,21.2\n")

    monkeypatch.setattr(requests, "get", fake_get)
    source = GoogleSheetsMetadataSource(
        "https://docs.google.com/spreadsheets/d/sheet-id/edit",
        sheet_name="Missing Sheet",
        file_name_column="File",
        coordinate_mapping={"Energy": "hv"},
    )

    with pytest.raises(
        ValueError,
        match=r"Missing Sheet.*Available worksheets: 'Linked Sheet', 'Other Sheet'",
    ):
        source.refresh()
    assert params == [
        {
            "tqx": "out:csv",
            "headers": "0",
            "sheet": "Missing Sheet",
            "range": "1:1",
        }
    ]


def test_google_source_rejects_missing_linked_gid(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        requests,
        "get",
        lambda *_args, **_kwargs: _FakeResponse(
            _google_sheet_list(("First Sheet", 0), ("Second Sheet", 456))
        ),
    )
    source = GoogleSheetsMetadataSource(
        "https://docs.google.com/spreadsheets/d/sheet-id/edit#gid=123",
        file_name_column="File",
        coordinate_mapping={"Energy": "hv"},
    )

    with pytest.raises(ValueError, match=r"gid 123.*was not found"):
        source.refresh()


def test_google_source_decodes_worksheet_names(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        requests,
        "get",
        lambda *_args, **_kwargs: _FakeResponse(
            r'items.push({name: "A \x26 B \"quoted\"", '
            r'pageUrl: "", gid: "123", initialSheet: false});'
        ),
    )
    source = GoogleSheetsMetadataSource(
        "https://docs.google.com/spreadsheets/d/sheet-id/edit"
    )

    assert source.get_sheet_names() == ['A & B "quoted"']


def test_google_source_uses_first_worksheet_without_linked_gid(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    params: list[dict[str, str]] = []

    def fake_get(url: str, **kwargs) -> _FakeResponse:
        if url.endswith("/htmlview"):
            return _FakeResponse(
                _google_sheet_list(("First Sheet", 321), ("Second Sheet", 654))
            )
        params.append(kwargs["params"])
        return _FakeResponse("File,Energy\n1,21.2\n")

    monkeypatch.setattr(requests, "get", fake_get)
    source = GoogleSheetsMetadataSource(
        "https://docs.google.com/spreadsheets/d/sheet-id/edit",
        file_name_column="File",
        coordinate_mapping={"Energy": "hv"},
    )

    assert source._metadata_for_file_number(1) is not None
    assert params == [{"tqx": "out:csv", "headers": "1"}]


def test_google_source_uses_one_based_spreadsheet_row_range(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    params: list[dict[str, str]] = []
    content_reads = 0

    def fake_get(url: str, **kwargs) -> _FakeResponse:
        nonlocal content_reads
        if url.endswith("/htmlview"):
            return _FakeResponse(_google_sheet_list(("Other Sheet", 456)))
        request_params = kwargs["params"]
        params.append(request_params)
        if request_params["range"] == "20:27":
            content_reads += 1
            return _FakeResponse(f"f_0001,{content_reads * 20.0}\n")
        return _FakeResponse("File,Energy\n")

    monkeypatch.setattr(requests, "get", fake_get)
    source = GoogleSheetsMetadataSource(
        "https://docs.google.com/spreadsheets/d/sheet-id/edit",
        sheet_name="Other Sheet",
        row_range=(20, 27),
        file_name_column="File",
        coordinate_mapping={"Energy": "hv"},
    )

    values = source._metadata_for_file_number(1)
    assert values is not None
    assert values.coordinate_values == {"hv": 20.0}

    updated_values = source._metadata_for_file_number(1)
    assert updated_values is not None
    assert updated_values.coordinate_values == {"hv": 40.0}
    assert {
        (request.get("range"), request.get("sheet"), request.get("gid"))
        for request in params
    } == {
        ("1:1", "Other Sheet", None),
        ("20:27", "Other Sheet", None),
        ("20:27", None, "456"),
    }


def test_google_source_initial_requests_run_concurrently(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    request_barrier = threading.Barrier(3)

    def fake_get(url: str, **kwargs) -> _FakeResponse:
        request_barrier.wait(timeout=5)
        if url.endswith("/htmlview"):
            return _FakeResponse(_google_sheet_list(("Sheet1", 123)))
        if kwargs["params"]["range"] == "1:1":
            return _FakeResponse("File,Energy\n")
        return _FakeResponse("f_0001,21.2\n")

    monkeypatch.setattr(requests, "get", fake_get)
    source = GoogleSheetsMetadataSource(
        "https://docs.google.com/spreadsheets/d/sheet-id/edit#gid=123",
        row_range=(20, 27),
        file_name_column="File",
        coordinate_mapping={"Energy": "hv"},
    )

    values = source._metadata_for_file_number(1)
    assert values is not None
    assert values.coordinate_values == {"hv": 21.2}


@pytest.mark.parametrize(
    ("status_code", "match"),
    [
        (401, "not publicly readable"),
        (403, "not publicly readable"),
        (429, "rate-limited"),
        (500, "HTTP 500"),
    ],
)
def test_google_source_http_errors(
    monkeypatch: pytest.MonkeyPatch, status_code: int, match: str
) -> None:
    monkeypatch.setattr(
        requests,
        "get",
        lambda *_args, **_kwargs: _FakeResponse(status_code=status_code),
    )
    source = GoogleSheetsMetadataSource(
        "https://docs.google.com/spreadsheets/d/sheet-id/edit",
        sheet_name="Sheet1",
        file_name_column="File",
        coordinate_mapping={"Energy": "hv"},
    )

    with pytest.raises(RuntimeError, match=match):
        source.refresh()


def test_google_source_transport_and_csv_errors(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = GoogleSheetsMetadataSource(
        "https://docs.google.com/spreadsheets/d/sheet-id/edit",
        sheet_name="Sheet1",
        file_name_column="File",
        coordinate_mapping={"Energy": "hv"},
    )

    def fail_request(*_args, **_kwargs):
        raise requests.Timeout("timeout")

    monkeypatch.setattr(requests, "get", fail_request)
    with pytest.raises(RuntimeError, match="request failed"):
        source.refresh()

    monkeypatch.setattr(
        requests,
        "get",
        lambda url, **_kwargs: _FakeResponse(
            _google_sheet_list(("Sheet1", 0))
            if url.endswith("/htmlview")
            else 'File,Energy\n1,"unterminated\n'
        ),
    )
    with pytest.raises(RuntimeError, match="invalid CSV"):
        source.refresh()


def test_spreadsheet_scalar_parsing() -> None:
    assert metadata_core._parse_spreadsheet_value("0") == 0
    assert metadata_core._parse_spreadsheet_value("-12") == -12
    assert metadata_core._parse_spreadsheet_value("2.5") == 2.5
    assert metadata_core._parse_spreadsheet_value("1e3") == 1000.0
    assert metadata_core._parse_spreadsheet_value("TRUE") is True
    assert metadata_core._parse_spreadsheet_value("FALSE") is False
    assert metadata_core._parse_spreadsheet_value("001") == "001"
    assert metadata_core._parse_spreadsheet_value("2026-07-21") == "2026-07-21"
